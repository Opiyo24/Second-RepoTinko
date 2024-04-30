import openpyxl

# Load the workbook
# Create workbook object
wb = openpyxl.load_workbook('balance.xlsx')

#print all sheet names
# print(wb.sheetnames)

#select particular sheetname
ws = wb['Score']
# print(ws)

ws1 = wb['Sheet1']
# print(ws1)

#Create a anew worksheet
wb.create_sheet('NewSheet')

#Save
wb.save("balance.xlsx")
print(wb.sheetnames)